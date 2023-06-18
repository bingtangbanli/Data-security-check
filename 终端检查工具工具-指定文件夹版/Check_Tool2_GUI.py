import os
import re
from openpyxl import load_workbook
from docx import Document
import PySimpleGUI as sg
import threading

# 全局锁对象，用于保证线程安全
lock = threading.Lock()
stop_scan = False
scan_results = []  # 创建一个空的扫描结果列表
problem_files = []  # 创建一个空的问题文件列表
# 获取指定文件夹下指定扩展名的文件列表
def get_files_by_extension(folder_path, extensions):
    files = []
    for root, _, filenames in os.walk(folder_path):
        for filename in filenames:
            _, ext = os.path.splitext(filename)
            if ext.lower() in extensions:
                files.append(os.path.join(root, filename))
    return files

# 处理 Excel 文件
def process_excel_file(file_path, phone_pattern, id_pattern):
    try:
        workbook = load_workbook(file_path)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(values_only=True):
                for cell_value in row:
                    if cell_value:
                        cell_value = str(cell_value)
                        phone_matches = re.findall(phone_pattern, cell_value)
                        id_matches = re.findall(id_pattern, cell_value)
                        if phone_matches or id_matches:
                            return True
    except Exception as e:
        pass
    return False

# 处理 Word 文件
def process_word_file(file_path, phone_pattern, id_pattern):
    try:
        doc = Document(file_path)
        for paragraph in doc.paragraphs:
            text = paragraph.text
            phone_matches = re.findall(phone_pattern, text)
            id_matches = re.findall(id_pattern, text)
            if phone_matches or id_matches:
                return True
    except Exception as e:
        pass
    return False

# 处理文本文件
def process_txt_file(file_path, phone_pattern, id_pattern):
    try:
        with open(file_path, 'r') as file:
            content = file.read()
            phone_matches = re.findall(phone_pattern, content)
            id_matches = re.findall(id_pattern, content)
            if phone_matches or id_matches:
                return True
    except Exception as e:
        pass
    return False

def is_file_in_use(file_path):
    try:
        with open(file_path, 'r'):
            return False
    except IOError:
        return True

def print_progress(file_path, has_match, window):
    result = "存在问题" if has_match else "不存在问题"
    with lock:
        print(f"扫描文件: {file_path}，结果: {result}")
        window["-RESULT-"].print(f"扫描文件: {file_path}, 结果: {result}")
        if has_match:
            problem_files.append(file_path)  # 将有问题的文件添加到问题文件列表中

def scan_files(folder_path, phone_regex, id_regex, result_file, window):
    global stop_scan
    stop_scan = False

    excel_files = get_files_by_extension(folder_path, ['.xlsx', '.xls'])
    word_files = get_files_by_extension(folder_path, ['.docx', '.doc'])
    txt_files = get_files_by_extension(folder_path, ['.txt'])

    results = []

    for file_path in excel_files:
        if stop_scan:
            break

        if is_file_in_use(file_path):
            print(f"文件被占用: {file_path}")
            window.write_event_value('-PROGRESS-', f"文件被占用: {file_path}")
            continue

        has_match = process_excel_file(file_path, phone_regex, id_regex)
        if has_match:
            results.append(file_path)
        print_progress(file_path, has_match, window)
        scan_results.append(file_path)
        window['-RESULT-'].update('\n'.join(scan_results))

    for file_path in word_files:
        if stop_scan:
            break

        if is_file_in_use(file_path):
            print(f"文件被占用: {file_path}")
            window.write_event_value('-PROGRESS-', f"文件被占用: {file_path}")
            continue

        has_match = process_word_file(file_path, phone_regex, id_regex)
        if has_match:
            results.append(file_path)
        print_progress(file_path, has_match, window)
        scan_results.append(file_path)
        window['-RESULT-'].update('\n'.join(scan_results))

    for file_path in txt_files:
        if stop_scan:
            break

        if is_file_in_use(file_path):
            print(f"文件被占用: {file_path}")
            window.write_event_value('-PROGRESS-', f"文件被占用: {file_path}")
            continue

        has_match = process_txt_file(file_path, phone_regex, id_regex)
        if has_match:
            results.append(file_path)
        print_progress(file_path, has_match, window)
        scan_results.append(file_path)
        window['-RESULT-'].update('\n'.join(scan_results))

    with open(result_file, 'w') as file:
        for result in results:
            file.write(f"{result}\n")

    window.write_event_value('-SCAN_COMPLETE-', results)

    return results

def main():
    sg.theme("DefaultNoMoreNagging")

    layout = [
        [sg.Text("选择文件夹路径: "), sg.InputText(key="-FOLDER_PATH-"), sg.FolderBrowse(key="-BROWSE-")],
        [sg.Text("手机号正则表达式: "), sg.InputText(key="-PHONE_REGEX-", default_text=r'^1[3-9]\d{9}$')],
        [sg.Text("身份证号正则表达式: "), sg.InputText(key="-ID_REGEX-", default_text=r'\b\d{17}[\dXx]\b')],
        [sg.Button("开始扫描", key="-START_SCAN-"), sg.Button("停止扫描", key="-STOP_SCAN-", disabled=True)],
        [sg.Text("扫描结果: ")],
        [sg.Multiline(size=(60, 10), key="-RESULT-", disabled=True)],
        [sg.Text(size=(60,1), key='-STATUS-', justification='center')]
    ]

    window = sg.Window("Check_Tool2_GUI-指定文件夹版（by 冰糖葫芦）", layout)

    while True:
        event, values = window.read()

        if event == sg.WINDOW_CLOSED:
            break
        elif event == "-START_SCAN-":
            folder_path = values["-FOLDER_PATH-"]
            phone_regex = values["-PHONE_REGEX-"]
            id_regex = values["-ID_REGEX-"]

            if folder_path:
                window["-START_SCAN-"].update(disabled=True)
                window["-STOP_SCAN-"].update(disabled=False)
                result_file = "scan_results.txt"  # 指定结果文件路径
                scan_thread = threading.Thread(target=scan_files, args=(folder_path, phone_regex, id_regex, result_file, window))
                scan_thread.start()
        elif event == "-STOP_SCAN-":
            stop_scan = True
            window["-STOP_SCAN-"].update(disabled=True)
        elif event == '-PROGRESS-':
            window['-STATUS-'].update(values[event])
        elif event == '-SCAN_COMPLETE-':
            window["-START_SCAN-"].update(disabled=False)
            window["-STOP_SCAN-"].update(disabled=True)
            window['-STATUS-'].update('扫描结束')
            sg.Popup("扫描完成！")
            sg.Popup("扫描完成！问题文件已保存到 scan_results.txt 文件中！")
            problem_files_str = "\n".join(problem_files)  # 将问题文件列表转换为字符串
            sg.Popup(f"存在问题的文件：\n{problem_files_str}")  # 显示存在问题的文件列表

    window.close()

if __name__ == "__main__":
    main()
