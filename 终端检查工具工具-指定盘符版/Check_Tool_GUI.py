import os
import re
from openpyxl import load_workbook
from docx import Document
import PySimpleGUI as sg
import threading

# 全局锁对象，用于线程安全
lock = threading.Lock()
stop_scan = False
scan_results = []  # 创建一个空列表存储扫描结果
problem_files = []  # 创建一个空列表存储问题文件

# 获取系统上的驱动器列表
def get_drives():
    drives = []
    for drive in range(ord('A'), ord('Z') + 1):
        drive = chr(drive) + ':\\'
        if os.path.exists(drive):
            drives.append(drive)
    return drives

# 获取指定驱动器上与扩展名匹配的文件列表
def get_files_by_drive(drive, extensions):
    files = []
    for root, _, filenames in os.walk(drive):
        for filename in filenames:
            _, ext = os.path.splitext(filename)
            if ext.lower() in extensions:
                files.append(os.path.join(root, filename))
    return files

# 处理Excel文件
def process_excel_file(file_path, phone_pattern, id_pattern):
    try:
        workbook = load_workbook(file_path)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(values_only=True):
                for cell_value in row:
                    if cell_value:
                        cell_value = str(cell_value)
                        phone_matches = re.findall(phone_pattern, cell_value)
                        id_matches = re.findall(id_pattern, cell_value)
                        if phone_matches or id_matches:
                            return True
    except Exception as e:
        pass
    return False

# 处理Word文件
def process_word_file(file_path, phone_pattern, id_pattern):
    try:
        doc = Document(file_path)
        for paragraph in doc.paragraphs:
            text = paragraph.text
            phone_matches = re.findall(phone_pattern, text)
            id_matches = re.findall(id_pattern, text)
            if phone_matches or id_matches:
                return True
    except Exception as e:
        pass
    return False

# 检查文件是否正在使用中
def is_file_in_use(file_path):
    try:
        with open(file_path, 'r'):
            return False
    except IOError:
        return True

# 打印扫描进度
def print_progress(file_path, has_match, window):
    result = "有问题" if has_match else "没有问题"
    with lock:
        print(f"正在扫描文件: {file_path}，结果: {result}")
        window["-RESULT-"].print(f"正在扫描文件: {file_path}，结果: {result}")
        if has_match:
            problem_files.append(file_path)

# 扫描文件
def scan_files(drives, phone_regex, id_regex, result_file, window):
    global stop_scan
    stop_scan = False

    extensions = ['.xlsx', '.xls', '.docx', '.doc', '.txt']

    results = []

    for drive in drives:
        if stop_scan:
            break

        for file_path in get_files_by_drive(drive, extensions):
            if stop_scan:
                break

            if is_file_in_use(file_path):
                print(f"文件正在使用中: {file_path}")
                window.write_event_value('-PROGRESS-', f"文件正在使用中: {file_path}")
                continue

            if file_path.endswith(('.xlsx', '.xls')):
                has_match = process_excel_file(file_path, phone_regex, id_regex)
            elif file_path.endswith(('.docx', '.doc')):
                has_match = process_word_file(file_path, phone_regex, id_regex)
            else:
                continue

            if has_match:
                results.append(file_path)

            print_progress(file_path, has_match, window)
            scan_results.append(file_path)
            window['-RESULT-'].update('\n'.join(scan_results))

    with open(result_file, 'w') as file:
        for result in results:
            file.write(f"{result}\n")

    window.write_event_value('-SCAN_COMPLETE-', results)

    return results

def main():
    sg.theme("DefaultNoMoreNagging")

    drives = get_drives()  # 获取系统上的驱动器列表

    layout = [
        [sg.Text("选择要扫描的驱动器:")],
        [sg.Listbox(drives, size=(20, min(len(drives), 10)), key='-DRIVES-', enable_events=True, select_mode='extended')],
        [sg.Text("手机号正则表达式模式:"), sg.InputText(key="-PHONE_REGEX-", default_text=r'^1[3-9]\d{9}$')],
        [sg.Text("身份证号正则表达式模式:"), sg.InputText(key="-ID_REGEX-", default_text=r'\b\d{17}[\dXx]\b')],
        [sg.Button("开始扫描", key="-START_SCAN-"), sg.Button("暂停扫描", key="-PAUSE_SCAN-", disabled=True)],
        [sg.Text("扫描结果:")],
        [sg.Multiline(size=(60, 10), key="-RESULT-", disabled=True)],
        [sg.Text(size=(60, 1), key='-STATUS-', justification='center')]
    ]

    window = sg.Window("Check_Tool_GUI-指定盘符版(by 冰糖葫芦)", layout)

    while True:
        event, values = window.read()

        if event == sg.WINDOW_CLOSED:
            break
        elif event == "-START_SCAN-":
            drives_selected = values['-DRIVES-']
            if not drives_selected:
                sg.Popup("请先选择要扫描的驱动器!")
                continue

            phone_regex = values["-PHONE_REGEX-"]
            id_regex = values["-ID_REGEX-"]

            window["-START_SCAN-"].update(disabled=True)
            window["-PAUSE_SCAN-"].update(disabled=False)
            result_file = "scan_all_results.txt"  # 指定结果文件路径
            scan_thread = threading.Thread(target=scan_files, args=(drives_selected, phone_regex, id_regex, result_file, window))
            scan_thread.start()
        elif event == "-PAUSE_SCAN-":
            stop_scan = True
            window["-PAUSE_SCAN-"].update(disabled=True)
        elif event == '-PROGRESS-':
            window['-STATUS-'].update(values[event])
        elif event == '-SCAN_COMPLETE-':
            window["-START_SCAN-"].update(disabled=False)
            window["-PAUSE_SCAN-"].update(disabled=True)
            window['-STATUS-'].update('扫描完成')
            sg.Popup("扫描完成！问题文件已保存到 scan_all_results.txt 文件中！")
            sg.Popup("扫描完成!")
            problem_files_str = "\n".join(problem_files)
            sg.popup_scrolled(problem_files_str, title='问题文件列表')

    window.close()

if __name__ == "__main__":
    main()
